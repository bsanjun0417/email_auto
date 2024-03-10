import openpyxl
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.encoders import encode_base64 #한글깨짐방지
from email.header import Header
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart #파일첨부
from email.utils import formatdate
from email import encoders
import sqlite3



conn = sqlite3.connect('main.db')
cursor = conn.cursor()
cursor.execute('select * from email_data;')
data_list = cursor.fetchall()


nmail_id= data_list[0][1]
nmail_pw=data_list[0][2]

gmail_id = data_list[1][1]
gmail_pw = data_list[1][2]



def excelplay(email_select,file_path):
    test(email_select,file_path)




def test(email_select,file_path):
    print("file",file_path)
    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook['이메일 자동전송']
    row_index = 1 # 예시로 2행

    column_index = ['A','B','C','D']
    title = ["이메일","제목","내용","파일경로"]
    title_check=0
    for i in range(0,4):
        print(title_check)
        title_list = sheet[column_index[i] + "1"].value
        if title_list == title[i]:
            title_check += 1
        elif title_list != title[i]:
            title_check = 0
            break

    if title_check == 4:
        print(title_check,"다음함수실행")
        excel_test(email_select,file_path)


       # 엑셀 파일 닫기
    workbook.close()





excel_data =[]
def excel_test(email_select,file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['이메일 자동전송']
    last = sheet.max_row + 1
    print("last test",last)
    for i in range(2,last):
        cell_value1 = sheet['A'+str(i)].value
        cell_value2 = sheet['B'+str(i)].value
        cell_value3 = sheet['C'+str(i)].value
        celldata = sheet['D'+str(i)].value

        if os.path.exists(celldata): #해당 경로에 파일이 있는지 확인해줌 exits
            cell_value4 = celldata.replace('\\','\\\\')  
            print(cell_value4,"data")
        else:
            cell_value4 = 'none'


        insert_data=[cell_value1,cell_value2,cell_value3,cell_value4]
        excel_data.append(insert_data)
        #배열로 insert_data계속 넣기
        #함수 실행후 파일경로에 파일이 없으면 none
    send_xlsx(email_select,file_path)
 


def send_xlsx(email_select,file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['이메일 자동전송']
    last = sheet.max_row - 1
    print("send _ test:",excel_data)
    if email_select == 'naver':
        print("최종")
        for i in range(0,last):
            naver_send(excel_data[i][0],excel_data[i][1],excel_data[i][2],excel_data[i][3])

    elif email_select == 'google':
        for i in range(0,last):
            gmail_send(excel_data[i][0],excel_data[i][1],excel_data[i][2],excel_data[i][3])
                

def naver_send(to_email,title,body,path):
    msg = MIMEMultipart()
    msg['Subject'] = title
    msg['From'] = nmail_id
    msg['To']= to_email
    
    msg.attach(MIMEText(body,"plain", 'utf-8'))
    with open(path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
                    "Content-Disposition",
                    f"attachment; filename= {os.path.basename(path)}",
        )
        msg.attach(part)
    smtp = smtplib.SMTP("smtp.naver.com",587)
    smtp.starttls()

    smtp.login(user=nmail_id,password=nmail_pw)
    smtp.sendmail(nmail_id,to_email,msg.as_string())
    smtp.close()




def gmail_send(to_email,title,body,path):
    msg = MIMEMultipart()
    msg['Subject'] = title
    msg['From'] = gmail_id
    msg['To']= to_email
    
    msg.attach(MIMEText(body,"plain", 'utf-8'))
    with open(path, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
                    "Content-Disposition",
                    f"attachment; filename= {os.path.basename(path)}",
        )
        msg.attach(part)
    smtp = smtplib.SMTP("smtp.naver.com",587)
    smtp.starttls()

    smtp.login(user=gmail_id,password=gmail_pw)
    smtp.sendmail(gmail_id,to_email,msg.as_string())
    smtp.close()








